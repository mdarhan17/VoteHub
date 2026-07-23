from io import BytesIO
from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    flash,
    send_file
)
from app.extensions import mysql
from app.utils.decorators import admin_required
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

result_bp = Blueprint(
    "result",
    __name__,
    url_prefix="/admin/results"
)


def synchronize_election_status(election_id=None):
    cur = mysql.connection.cursor()

    if election_id:
        cur.execute("""
            UPDATE elections
            SET status='completed'
            WHERE election_id=%s
              AND status='active'
              AND end_datetime <= NOW()
        """, (election_id,))
    else:
        cur.execute("""
            UPDATE elections
            SET status='completed'
            WHERE status='active'
              AND end_datetime <= NOW()
        """)

    mysql.connection.commit()
    cur.close()


def get_result_data(election_id):
    synchronize_election_status(election_id)

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT *
        FROM elections
        WHERE election_id=%s
    """, (election_id,))
    election = cur.fetchone()

    if not election:
        cur.close()
        return None, [], 0, [], False

    cur.execute("""
        SELECT
            c.candidate_id,
            c.full_name,
            c.party_name,
            c.photo,
            c.symbol,
            c.description,
            COUNT(v.vote_id) AS total_votes
        FROM election_candidates ec
        INNER JOIN candidates c
            ON ec.candidate_id = c.candidate_id
        LEFT JOIN votes v
            ON v.candidate_id = c.candidate_id
           AND v.election_id = ec.election_id
        WHERE ec.election_id=%s
        GROUP BY
            c.candidate_id,
            c.full_name,
            c.party_name,
            c.photo,
            c.symbol,
            c.description
        ORDER BY
            total_votes DESC,
            c.full_name ASC
    """, (election_id,))
    results = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM votes
        WHERE election_id=%s
    """, (election_id,))
    total_votes = cur.fetchone()["total"]

    cur.close()

    winners = []
    is_tie = False

    if results and total_votes > 0:
        highest_votes = results[0]["total_votes"]

        winners = [
            candidate
            for candidate in results
            if candidate["total_votes"] == highest_votes
        ]

        is_tie = len(winners) > 1

    return election, results, total_votes, winners, is_tie


@result_bp.route("/")
@admin_required
def results():
    synchronize_election_status()

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT
            e.*,
            COUNT(DISTINCT v.vote_id) AS total_votes,
            COUNT(DISTINCT ec.candidate_id) AS candidate_count
        FROM elections e
        LEFT JOIN votes v
            ON e.election_id = v.election_id
        LEFT JOIN election_candidates ec
            ON e.election_id = ec.election_id
        GROUP BY e.election_id
        ORDER BY e.election_id DESC
    """)

    election_list = cur.fetchall()
    cur.close()

    return render_template(
        "admin/result_dashboard.html",
        elections=election_list
    )


@result_bp.route("/election/<int:election_id>")
@admin_required
def election_result(election_id):
    (
        election,
        results,
        total_votes,
        winners,
        is_tie
    ) = get_result_data(election_id)

    if not election:
        flash("Election not found.", "danger")
        return redirect(url_for("result.results"))

    return render_template(
        "admin/election_result.html",
        election=election,
        results=results,
        total_votes=total_votes,
        winners=winners,
        is_tie=is_tie
    )


@result_bp.route("/publish/<int:election_id>")
@admin_required
def publish_result(election_id):
    (
        election,
        results,
        total_votes,
        winners,
        is_tie
    ) = get_result_data(election_id)

    if not election:
        flash("Election not found.", "danger")
        return redirect(url_for("result.results"))

    if election["status"] != "completed":
        flash(
            "Complete the election before publishing its result.",
            "warning"
        )
        return redirect(url_for("result.results"))

    if total_votes <= 0:
        flash(
            "The result cannot be published because no votes were cast.",
            "warning"
        )
        return redirect(url_for("result.results"))

    cur = mysql.connection.cursor()

    cur.execute("""
        UPDATE elections
        SET result_published=1
        WHERE election_id=%s
    """, (election_id,))

    mysql.connection.commit()
    cur.close()

    flash("Result published successfully.", "success")
    return redirect(url_for("result.results"))


@result_bp.route("/unpublish/<int:election_id>")
@admin_required
def unpublish_result(election_id):
    cur = mysql.connection.cursor()

    cur.execute("""
        UPDATE elections
        SET result_published=0
        WHERE election_id=%s
    """, (election_id,))

    mysql.connection.commit()
    cur.close()

    flash("Result unpublished successfully.", "warning")
    return redirect(url_for("result.results"))


@result_bp.route("/export/pdf/<int:election_id>")
@admin_required
def export_pdf(election_id):
    (
        election,
        results,
        total_votes,
        winners,
        is_tie
    ) = get_result_data(election_id)

    if not election:
        flash("Election not found.", "danger")
        return redirect(url_for("result.results"))

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    def create_header():
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(
            50,
            height - 55,
            "VoteHub - Election Result Report"
        )

        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(
            50,
            height - 88,
            f"Election: {election['title']}"
        )

        pdf.setFont("Helvetica", 10)
        pdf.drawString(
            50,
            height - 108,
            f"Status: {str(election['status']).title()}"
        )

        pdf.drawString(
            50,
            height - 125,
            f"Total Votes Cast: {total_votes}"
        )

        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(50, height - 160, "Rank")
        pdf.drawString(100, height - 160, "Candidate")
        pdf.drawString(300, height - 160, "Party")
        pdf.drawString(485, height - 160, "Votes")

        pdf.line(
            50,
            height - 168,
            width - 50,
            height - 168
        )

        return height - 190

    y = create_header()
    pdf.setFont("Helvetica", 10)

    for index, item in enumerate(results, start=1):
        if y < 90:
            pdf.showPage()
            y = create_header()
            pdf.setFont("Helvetica", 10)

        candidate_name = str(item["full_name"] or "")[:30]
        party_name = str(
            item["party_name"] or "Independent"
        )[:25]

        pdf.drawString(50, y, str(index))
        pdf.drawString(100, y, candidate_name)
        pdf.drawString(300, y, party_name)
        pdf.drawString(
            485,
            y,
            str(item["total_votes"])
        )

        y -= 22

    if winners:
        if y < 120:
            pdf.showPage()
            y = height - 70

        y -= 15
        pdf.setFont("Helvetica-Bold", 13)

        if is_tie:
            pdf.drawString(50, y, "Result: Tie")
            y -= 22

            pdf.setFont("Helvetica", 11)

            for winner in winners:
                pdf.drawString(
                    50,
                    y,
                    (
                        f"{winner['full_name']} - "
                        f"{winner['total_votes']} out of "
                        f"{total_votes} total votes"
                    )
                )
                y -= 18
        else:
            winner = winners[0]

            pdf.drawString(
                50,
                y,
                f"Winner: {winner['full_name']}"
            )

            y -= 20
            pdf.setFont("Helvetica", 11)
            pdf.drawString(
                50,
                y,
                (
                    f"Votes Received: "
                    f"{winner['total_votes']} out of "
                    f"{total_votes} total votes"
                )
            )
    else:
        y -= 15
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(
            50,
            y,
            "No winner available because no votes were cast."
        )

    pdf.save()
    buffer.seek(0)

    safe_title = "".join(
        character
        if character.isalnum()
        else "_"
        for character in election["title"]
    ).strip("_")

    filename = (
        f"{safe_title or 'election'}_result_"
        f"{election_id}.pdf"
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


@result_bp.route("/export/excel/<int:election_id>")
@admin_required
def export_excel(election_id):
    (
        election,
        results,
        total_votes,
        winners,
        is_tie
    ) = get_result_data(election_id)

    if not election:
        flash("Election not found.", "danger")
        return redirect(url_for("result.results"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Election Result"

    title_fill = PatternFill(
        "solid",
        fgColor="1D4ED8"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="DBEAFE"
    )

    sheet.merge_cells("A1:D1")
    sheet["A1"] = "VoteHub - Election Result Report"
    sheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    sheet.append([])
    sheet.append(["Election", election["title"]])
    sheet.append([
        "Status",
        str(election["status"]).title()
    ])
    sheet.append(["Total Votes Cast", total_votes])
    sheet.append([])

    sheet.append([
        "Rank",
        "Candidate",
        "Party",
        "Total Votes"
    ])

    for cell in sheet[7]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )

    for index, item in enumerate(results, start=1):
        sheet.append([
            index,
            item["full_name"],
            item["party_name"] or "Independent",
            item["total_votes"]
        ])

    sheet.append([])

    if winners:
        if is_tie:
            sheet.append(["Result", "Tie"])

            for winner in winners:
                sheet.append([
                    "Joint Winner",
                    winner["full_name"],
                    winner["party_name"] or "Independent",
                    (
                        f"{winner['total_votes']} out of "
                        f"{total_votes}"
                    )
                ])
        else:
            winner = winners[0]

            sheet.append([
                "Winner",
                winner["full_name"],
                winner["party_name"] or "Independent",
                (
                    f"{winner['total_votes']} out of "
                    f"{total_votes}"
                )
            ])
    else:
        sheet.append([
            "Result",
            "No winner because no votes were cast."
        ])

    column_widths = {
        "A": 18,
        "B": 32,
        "C": 28,
        "D": 20
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center"
            )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    safe_title = "".join(
        character
        if character.isalnum()
        else "_"
        for character in election["title"]
    ).strip("_")

    filename = (
        f"{safe_title or 'election'}_result_"
        f"{election_id}.xlsx"
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
